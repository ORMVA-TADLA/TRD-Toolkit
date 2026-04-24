import os
import sys
import argparse
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import range_boundaries


# ==========================================
# 1. DATA EXTRACTION (I/O)
# ==========================================
def extract_table_from_excel(filepath, table_name, required_cols):
    """
    Opens an Excel file, locates a specific table, and extracts the required
    columns into a pandas DataFrame.
    """
    print(f"\nOpening '{filepath}' and searching for table '{table_name}'...")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except FileNotFoundError:
        print(f"\n[ERROR] The file '{filepath}' could not be found.")
        sys.exit(1)

    target_ws = None
    target_table = None
    for ws in wb.worksheets:
        if table_name in ws.tables:
            target_ws = ws
            target_table = ws.tables[table_name]
            break

    if not target_table:
        print(f"\n[ERROR] Could not find table '{table_name}'. Check your config.")
        sys.exit(1)

    print(f"Found Table: '{target_table.name}'. Reading columns...")

    min_col, min_row, max_col, max_row = range_boundaries(target_table.ref)

    headers = [
        target_ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)
    ]

    col_indices = {}
    for col_name in required_cols:
        try:
            col_indices[col_name] = headers.index(col_name)
        except ValueError:
            print(f"\n[ERROR] Missing column: '{col_name}'. Headers found: {headers}")
            sys.exit(1)

    data = []
    for r in range(min_row + 1, max_row + 1):
        row_data = {}
        for col_name, idx in col_indices.items():
            row_data[col_name] = target_ws.cell(row=r, column=min_col + idx).value
        data.append(row_data)

    return pd.DataFrame(data)


# ==========================================
# 2. BUSINESS LOGIC (MATH)
# ==========================================
def allocate_proportional_integers(group_df, val_col, target_col, output_col):
    """
    Takes a grouped subset of data, scales the values proportionally to hit
    a target sum, and uses the largest remainder method to ensure integer outputs.
    """
    target_sum = float(group_df[target_col].iloc[0])

    clean_vals = pd.to_numeric(group_df[val_col], errors="coerce").fillna(0)
    orig_nums = clean_vals.values

    orig_sum = np.sum(orig_nums)

    if orig_sum == 0:
        group_df[output_col] = 0
        return group_df

    exact_vals = orig_nums * (target_sum / orig_sum)
    base_ints = np.floor(exact_vals).astype(int)
    remainders = exact_vals - base_ints
    shortfall = int(target_sum - np.sum(base_ints))

    if shortfall > 0:
        largest_indices = np.argsort(remainders)[-shortfall:]
        for idx in largest_indices:
            base_ints[idx] += 1

    group_df[output_col] = base_ints

    return group_df


# ==========================================
# 3. ORCHESTRATION (MAIN WORKFLOW)
# ==========================================
def process_data(
    filepath, target_table_name, id_col, group_col, val_col, target_col, output_col
):
    """
    Coordinates the extraction, calculation, and exporting of the data.
    """
    required_cols = [id_col, group_col, val_col, target_col]

    df = extract_table_from_excel(filepath, target_table_name, required_cols)

    print("Calculating final numbers...")

    result_df = df.groupby(group_col, group_keys=False).apply(
        lambda g: allocate_proportional_integers(g, val_col, target_col, output_col)
    )

    # [CHANGED]: Swapped group_col and id_col so the ID is the 2nd column
    final_column_order = [group_col, id_col, val_col, target_col, output_col]
    result_df = result_df[final_column_order]

    base_name, _ = os.path.splitext(filepath)
    output_csv_path = f"{base_name}_output.csv"

    result_df.to_csv(output_csv_path, index=False)

    print(f"\n[SUCCESS] Calculations complete. Data saved to '{output_csv_path}'")


# ==========================================
# 4. CLI ENTRY POINT
# ==========================================
if __name__ == "__main__":
    # ------------------------------------------
    # CONFIGURATION: Set your fixed target names
    # ------------------------------------------
    TABLE_NAME = "PC"
    ID_COL_NAME = "Code Parcelle"
    GROUP_COL_NAME = "Ref"
    VAL_COL_NAME = "Hours static"
    TARGET_COL_NAME = "Target Sum"
    OUTPUT_COL_NAME = "Hours final"
    # ------------------------------------------

    parser = argparse.ArgumentParser(
        description="Reads an Excel table, scales numbers to integers, and exports to CSV.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog="Example usage: python scale_table.py my_data.xlsx",
    )

    parser.add_argument(
        "filepath",
        nargs="?",
        help="Path to the .xlsx file (if omitted, you will be prompted)",
    )

    args = parser.parse_args()

    target_filepath = args.filepath
    if not target_filepath:
        target_filepath = input("Please enter the path to your Excel file: ").strip()

        if not target_filepath:
            print("\n[ERROR] No file path provided. Exiting.")
            sys.exit(1)

    process_data(
        filepath=target_filepath,
        target_table_name=TABLE_NAME,
        id_col=ID_COL_NAME,
        group_col=GROUP_COL_NAME,
        val_col=VAL_COL_NAME,
        target_col=TARGET_COL_NAME,
        output_col=OUTPUT_COL_NAME,
    )
