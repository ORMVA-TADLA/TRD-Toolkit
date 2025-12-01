from access_parser import AccessParser
from collections import defaultdict
import pandas as pd
from datetime import datetime, timedelta
import csv
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os


def extract_ter(refter):
    parts = refter.split("T")
    return "T" + parts[1]


def calculate_hours_difference(start_date, end_date):
    """
    Calculates the difference in hours between two datetime objects.
    """
    time_difference = end_date - start_date
    difference_in_hours = time_difference.total_seconds() / 3600
    return difference_in_hours


def sum_chunks(data_list, chunk_sizes):
    """
    Splits a list into chunks of specified sizes and calculates the average of each chunk.
    """
    result_list = []
    start_index = 0
    pattern_index = 0

    while start_index < len(data_list):
        chunk_size = chunk_sizes[pattern_index % len(chunk_sizes)]
        end_index = start_index + chunk_size

        # Ensures the last chunk doesn't go past the end of the list
        chunk = data_list[start_index:end_index]
        result_list.append(sum(chunk) / chunk_size)

        start_index = end_index
        pattern_index += 1

    return result_list


def read_mdb_file(file_path):
    """
    Read MDB file using access-parser library.
    Works for both .mdb and .accdb files.
    """
    # Initialize parser
    db = AccessParser(file_path)

    # Get all table names
    tables = db.catalog

    # Tables are stored as defaultdict(list) -- table[column][row_index]
    result = {}
    for table_name in tables:
        table_data = db.parse_table(table_name)
        result[table_name] = table_data

    return result


def table_to_csv(data, table_name, output_file):
    """
    Save table data to a CSV file.
    """
    df = pd.DataFrame(data[table_name])
    df.to_csv(output_file, index=False)
    print(f"Data from table '{table_name}' saved to {output_file}")
    return output_file


def table_to_xlsx(data, table_name, output_file):
    """
    Save table data to an XLS file.
    """
    df = pd.DataFrame(data[table_name])
    df.to_excel(output_file, index=False)
    print(f"Data from table '{table_name}' saved to {output_file}")
    return output_file


def mv1_to_dict(db):
    """
    Convert MV1 table data to a nested dictionary grouped by "refsec" and "refter".
    """
    # AGRICULTEUR
    # CodeClient
    # npagr
    # convert every row to dictionary, group by "refsec", then "refter", ignore rows with null "refsec" or "refter"
    mv1_table = db["MV1"]
    agr_table = db["AGRICULTEUR"]
    # Convert table to list of rows (dictionaries)
    rows = []
    num_rows = len(next(iter(mv1_table.values())))
    for i in range(num_rows):
        row = {col: mv1_table[col][i] for col in mv1_table}
        # find npagr from agr_table using CodeClient
        code_client = row.get("CodeClient", None)
        npagr = None
        for j in range(len(next(iter(agr_table.values())))):
            if agr_table["CodeClient"][j] == code_client:
                npagr = agr_table["npagr"][j]
                break
        row["npagr"] = npagr
        rows.append(row)

    return rows


def groupe_dict(mv1_dict):
    """
    Convert MV1 table data to a nested dictionary grouped by "refsec" and "refter".
    """
    # convert every row to dictionary, group by "refsec", then "refter", ignore rows with null "refsec" or "refter"
    # Group by "refsec" and "refter"
    grouped = defaultdict(lambda: defaultdict(list))
    for row in mv1_dict:
        key1 = row.get("refsec", None)
        key2 = row.get("refter", None)
        if key1 and key2:
            grouped[key1][key2].append(row)
            
    # Sort the sec and ter keys based on the numeric part
    grouped_sorted = dict(sorted(grouped.items(), key=lambda item: int(
        ''.join(filter(str.isdigit, item[0])))))
    for sec in grouped_sorted:
        grouped_sorted[sec] = dict(sorted(grouped_sorted[sec].items(), key=lambda item: int(
            ''.join(filter(str.isdigit, item[0])))))

    return grouped_sorted


def mv1_to_mv2(mv1_dict, mv2_type="ALL"):
    # Initialize variables
    TRD_start_hour = None
    TRD_end_hour = None
    # Find the earliest date in "dato2" column
    for refsec, refsec_dict in mv1_dict.items():
        for refter, refter_dict in refsec_dict.items():
            for item in refter_dict:
                str_date = item.get("dato2", None)
                if str_date != None:
                    date = datetime.strptime(str_date, '%Y-%m-%d %H:%M:%S')
                    if TRD_start_hour == None or date < TRD_start_hour:
                        TRD_start_hour = date

    # Set TRD_start_hour to 18:00 of that day
    TRD_start_hour = TRD_start_hour + timedelta(hours=18)

    mv2 = {}
    for refsec, refsec_dict in mv1_dict.items():
        for refter, refter_dict in refsec_dict.items():
            for item in refter_dict:
                # print(item)
                record_type = item.get("typem", None)
                if mv2_type != "ALL" and record_type != mv2_type:
                    continue

                duration = item.get("dur", 0)
                if duration == 0:
                    continue

                sec = refsec
                ter = extract_ter(item.get("refter", None))

                debit = item.get("debit", 0)

                open_day = item.get("dato2", None)
                open_hour = item.get("heureo2", None)
                open_date = datetime.strptime(
                    open_day, '%Y-%m-%d %H:%M:%S') + timedelta(hours=int(open_hour))

                if TRD_end_hour is None:
                    TRD_end_hour = open_date + timedelta(hours=duration)
                if open_date + timedelta(hours=duration) > TRD_end_hour:
                    TRD_end_hour = open_date + timedelta(hours=duration)

                if refsec not in mv2:
                    mv2[refsec] = {}
                if ter not in mv2[refsec]:
                    mv2[refsec][ter] = {"hours_list": [],
                                        "hours_list_summed": [], "total_hours": 0}

                date_diff = calculate_hours_difference(
                    TRD_start_hour, open_date)
                for i in range(duration):
                    hour_index = int(date_diff + i)
                    if hour_index >= len(mv2[sec][ter]["hours_list"]):
                        second_list = [0] * (
                            hour_index - len(mv2[sec][ter]["hours_list"]) + 1
                        )
                        mv2[sec][ter]["hours_list"].extend(second_list)
                        mv2[sec][ter]["hours_list"][hour_index] += debit
                    else:
                        mv2[sec][ter]["hours_list"][hour_index] += debit

        # Sort the sec and ter keys based on the numeric part
    mv2 = dict(sorted(mv2.items(), key=lambda item: int(
        ''.join(filter(str.isdigit, item[0])))))
    for sec in mv2:
        mv2[sec] = dict(sorted(mv2[sec].items(), key=lambda item: int(
            ''.join(filter(str.isdigit, item[0])))))

    # size of chunks, 15H for night, 9H for day
    CHUNKS_PATTERN = [15, 9]
    for sec in mv2:
        for ter in mv2[sec]:
            summed_list = sum_chunks(
                mv2[sec][ter]["hours_list"], CHUNKS_PATTERN)
            for value in summed_list:
                if value == 0:
                    mv2[sec][ter]["hours_list_summed"].append("")
                else:
                    mv2[sec][ter]["hours_list_summed"].append(
                        round(value, 2))
            mv2[sec][ter]["total_hours"] = sum(
                mv2[sec][ter]["hours_list"]) / 20

    return mv2, TRD_start_hour, TRD_end_hour


def mv2_to_xlsx(mv2, TRD_start_hour, TRD_end_hour, excel_file_name):
    """
    Creates a new Excel file (MV2) from the processed data.

    Args:
        mv2 (dict): The dictionary containing the processed data.
        TRD_start (datetime): The earliest start date.
        directory (str): The directory where the new file will be saved.

    Returns:
        str: The full path of the created Excel file.
    """
    # Calculate the number of days in the TRD period
    TRD_days = (TRD_end_hour - TRD_start_hour).days

    # Create a new workbook
    workbook = Workbook()

    # Get the active worksheet
    sheet = workbook.active

    # define grand total row
    grand_total_hours = [0]+[0, 0] * TRD_days

    # Define the style for the border's sides
    thin_black = Side(border_style="thin", color="000000")
    # Define the border using the sides
    border_all = Border(
        left=thin_black, right=thin_black, top=thin_black, bottom=thin_black
    )
    # Define fill
    fill = PatternFill(start_color="DDDDDD",
                       end_color="DDDDDD", fill_type="solid")
    # Define different Alignment objects
    title_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    data_alignment = Alignment(horizontal="center", vertical="center")
    keys_alignment = Alignment(horizontal="left", vertical="center")

    for sec in mv2:
        # Add an empty row between different 'sec' groups
        sheet.append([])
        sheet.append(
            [
                "",
                "Sec",
                "Ter",
                "Duration"
            ]
            + ["N", "J"] * TRD_days
        )
        # formatting header row
        for cell in sheet[sheet.max_row]:
            # Skip the first cell
            if cell.column == 1:
                continue
            cell.font = Font(name="Book Antiqua", size=10, bold=True)
            cell.alignment = title_alignment
            cell.fill = fill
            cell.border = border_all

        sec_total_duration = 0
        for ter in mv2[sec]:
            ter_duration = sum(mv2[sec][ter]["hours_list"]) / 20
            # Update grand total hours
            grand_total_hours[0] += ter_duration
            for i in range(len(mv2[sec][ter]["hours_list_summed"])):
                if isinstance(mv2[sec][ter]["hours_list_summed"][i], (int, float)):
                    grand_total_hours[1 +
                                      i] += mv2[sec][ter]["hours_list_summed"][i]
            sec_total_duration += ter_duration
            sheet.append(
                [""] + [sec, ter, ter_duration] +
                mv2[sec][ter]["hours_list_summed"]
            )
            # formatting data row
            for cell in sheet[sheet.max_row]:
                # Skip the first cell
                if cell.column == 1:
                    continue
                cell.font = Font(name="Book Antiqua", size=10)
                cell.border = border_all
                if cell.column in [2, 3]:  # 'Sec' and 'Ter' columns
                    cell.alignment = keys_alignment
                else:
                    cell.alignment = data_alignment
                if cell.column in [2, 3, 4]:  # 'Sec', 'Ter', and 'Duration' columns
                    cell.font = Font(name="Book Antiqua", size=10, bold=True)

        # Add a total row for the current 'sec'
        total_row = [""] + [sec, "Total", sec_total_duration]
        for i in range(5, 5 + 2 * TRD_days):
            col_sum = sum(
                sheet.cell(row=row_idx, column=i).value or 0
                for row_idx in range(sheet.max_row - len(mv2[sec]) + 1, sheet.max_row + 1)
            )
            total_row.append(round(col_sum, 2) if col_sum != 0 else "")
        sheet.append(total_row)
        # formatting total row
        for cell in sheet[sheet.max_row]:
            # Skip the first cell
            if cell.column == 1:
                continue
            cell.font = Font(name="Book Antiqua", size=10, bold=True)
            cell.alignment = title_alignment
            cell.fill = fill
            cell.border = border_all

    # Add a grand total row at the end
    sheet.append([])
    grand_total_row = ["", "Grand Total", ""] + grand_total_hours
    sheet.append(grand_total_row)
    # formatting grand total row
    for cell in sheet[sheet.max_row]:
        # Skip the first cell
        if cell.column == 1:
            continue
        cell.font = Font(name="Book Antiqua", size=12, bold=True)
        # check if cell value is string
        if isinstance(cell.value, str):
            cell.alignment = keys_alignment
        else:
            cell.alignment = data_alignment

    # add credit note
    sheet.append([])
    sheet.append(
        ["", "Generated by MV2 Creator app - by Anas Asimi - 2025"])
    sheet["B" + str(sheet.max_row)
          ].font = Font(name="Book Antiqua", size=10, italic=True)
    sheet["B" + str(sheet.max_row)].alignment = Alignment(horizontal="left")

    # Save the workbook to a file
    workbook.save(excel_file_name)
    return excel_file_name


# Example usage
if __name__ == "__main__":
    # reading a mdb file
    file_path = "Distribution.mdb"
    db = read_mdb_file(file_path)

    # saving to CSV
    # table_to_csv(data, table, "mv1.csv")

    # saving to XLS
    # table_to_xlsx(data, table, "mv1.xlsx")

    # converting MV1 table to grouped dict
    mv1_dict = mv1_to_dict(db)
    grouped = groupe_dict(mv1_dict)

    # converting grouped dict to mv2
    # mv2, TRD_start_hour, TRD_end_hour = mv1_to_mv2(grouped)

    # saving mv2 to excel file
    # mv2_to_xlsx(mv2, TRD_start_hour, TRD_end_hour, f"MV2 {TRD_start_hour} {TRD_end_hour}.xlsx")
